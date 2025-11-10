from __future__ import annotations

import json
import logging
import os
from collections import defaultdict
from datetime import date, datetime, time
from pathlib import Path
from typing import List, Dict, Union, Optional, Sequence, Any

import win32com.client as win32
from PySide6.QtCore import QObject
from openpyxl import load_workbook, Workbook
from openpyxl.cell import Cell
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pywintypes import com_error

from app.classes import Invoice, DirectPayrollAdjustment
################################################################################

__all__ = ("ExcelInterface",)

# Determines if we're running on Windows, which supports PDF export
IS_WINDOWS = os.name == "nt"
# Conditional Formatting rules to highlight missing customer names
YELLOW_FILL = PatternFill(bgColor="FFFF00")
IS_BLANK_RULE = FormulaRule(
    formula=['AND(LEN($A3)>0, LEN(TRIM($C3&""))=0)'],
    fill=YELLOW_FILL,
    stopIfTrue=True,
)

THIN_GRAY = Side(style="thin", color="D9D9D9")
ROW_BORDER = Border(bottom=THIN_GRAY)

################################################################################
class ExcelInterface(QObject):

    def __init__(self, wb_path: Path, lookup_path: Path, output_dir: Optional[Path]) -> None:
        super().__init__()

        self._wb_path: Path = wb_path
        self._lookup_path: Path = lookup_path
        self.output_dir: Optional[Path] = (output_dir or Path(".")).resolve()
        self.out_wb_path: Optional[Path] = None

        self.workbook: Workbook = None  # type: ignore
        self.lookup_wb: Workbook = None  # type: ignore
        self._master_tech_sheet: Worksheet = None  # type: ignore

        self._lookup_table: Dict[int, str] = {}
        # Dict of technician Names mapped to a dict with 'invoices' and 'dpa' lists
        self._technician_dict: Dict[str, Dict[str, List[Union[Invoice, DirectPayrollAdjustment]]]] = defaultdict(lambda: {"invoices": [], "dpa": []})
        self._invoice_to_name: Dict[int, str] = {}

        self.current_row: int = 1
        self.master_current_row: int = 1

        self._load_lookup_table()
        self._connect_main_wb()
        self._prepare_data()
        self._assert_sheet_integrity()

################################################################################
    def _load_lookup_table(self) -> None:
        """Load an invoice ID to customer name lookup table from the specified path.

        This lookup table may be in either Excel (.xlsx) or JSON (.json) format.

        JSON is preferred for performance, as it loads much faster than Excel, and
        will be generated automatically on the first usage of an Excel lookup file.
        """

        logging.info(f"Loading lookup table...")

        # If we have an Excel file, parse it and convert to JSON for next time
        if self._lookup_path.suffix == ".xlsx":
            logging.info("Excel lookup file detected.")
            try:
                self.lookup_wb = load_workbook(self._lookup_path)
                self._parse_lookup_wb()
            except Exception as ex:
                logging.critical(f"Failed to load lookup master workbook at {self._lookup_path}: {ex}")
                return
        # Else if we have a JSON file, load it directly
        elif self._lookup_path.suffix == ".json":
            logging.info("JSON lookup file detected.")
            try:
                with open(self._lookup_path, "r", encoding="utf-8") as f:
                    self._lookup_table = json.load(f)
            except Exception as ex:
                logging.critical(f"Failed to load lookup master JSON at {self._wb_path}: {ex}")
                return
        # Otherwise, unsupported format
        else:
            logging.critical(f"Unsupported lookup file format: {self._lookup_path.suffix}")
            return

################################################################################
    def _parse_lookup_wb(self) -> None:
        """Parses an Excel workbook of the proper schema to build the invoice ID
        to customer name lookup table.
        """

        logging.info("Parsing lookup master workbook...")

        # "Sheet1" is the expected sheet name exported from ServiceTitan
        if "Sheet1" not in self.lookup_wb.sheetnames:
            logging.critical("Lookup workbook missing 'Sheet1'")
            return

        lookup_ws: Worksheet = self.lookup_wb["Sheet1"]  # type: ignore

        count = 0
        # Parse each row in the lookup worksheet
        for i, row in enumerate(lookup_ws.iter_rows(min_row=2)):  # Skip header
            try:
                invoice_id = int(row[0].value)
                customer_name = row[1].value
            # ValueError: invalid literal for int()
            # TypeError: NoneType cannot be converted to int
            except (ValueError, TypeError):
                logging.critical(f"Failed to parse lookup worksheet row {i + 1}: {row}")
                continue
            else:
                count += 1

            self._lookup_table[invoice_id] = customer_name

        logging.info(f"Found {count} invoices in lookup master workbook.")

        # Save out the lookup table as JSON for next time
        out_path = self.output_dir / "LookupTable.json"
        json.dump(self._lookup_table, open(out_path, "w"))

        # Load the lookup table from the new JSON file
        self._lookup_path = out_path
        self._load_lookup_table()

################################################################################
    def _connect_main_wb(self):
        """Connect to the main workbook and confirm the required sheets exist."""

        try:
            self.workbook = load_workbook(self._wb_path)
        except Exception as ex:
            logging.critical(f"Failed to load workbook at {self._wb_path}: {ex}")
            return

        if "Invoices" not in self.workbook.sheetnames:
            logging.critical("Invoices sheet not found")
            return
        self._inv = self.workbook["Invoices"]

        if "Direct Payroll Adjustments" not in self.workbook.sheetnames:
            logging.critical("Direct Payroll Adjustments sheet not found")
            return
        self._dpa = self.workbook["Direct Payroll Adjustments"]

        logging.info("Successfully connected to workbook.")

################################################################################
    @staticmethod
    def validate_header_row(ws, *, expected: Sequence[str], row: int = 1, sheet_label: str = "Sheet") -> None:
        """Validate that the header row of the given worksheet matches the expected values.

        This is primarily to ensure that the workbook being processed is of the expected format.
        """

        # Normalizer function to ensure consistent comparison
        def _norm(s: object) -> str:
            # Normalize: Handle None, trim, collapse spaces, lowercase
            text = "" if s is None else str(s)
            return " ".join(text.strip().split()).lower()

        # Read exactly len(expected) cells from the header row to avoid excess columns
        actual_cells = [ws.cell(row=row, column=i + 1).value for i in range(len(expected))]
        for i, (exp, act) in enumerate(zip(expected, actual_cells), start=1):
            if _norm(act) != _norm(exp):
                logging.critical(
                    f"{sheet_label} header mismatch at column {i} ({get_column_letter(i)}): "
                    f"expected '{exp}', found '{act}'"
                )
                return

################################################################################
    def _assert_sheet_integrity(self) -> None:
        """Validate that the required sheets have the expected headers."""

        self.validate_header_row(
            self._inv,
            expected=[
                "Technician", "Invoice Id", "Invoice", "Invoiced On", "Customer",
                "Total", "Split %", "Subtotal", "Cost", "Bonus", "Pay Adj.",
                "NC Total", "Net Serv. Vol.", "GP", "Business Unit",
            ],
            sheet_label="Invoices",
        )

        self.validate_header_row(
            self._dpa,
            expected=[
                "Technician", "Invoice Id", "Invoice", "Posted On",
                "Memo", "Amount",
            ],
            sheet_label="Direct Payroll Adjustments",
        )

################################################################################
    def _prepare_data(self) -> None:
        """Prepare and parse data from the main workbook into Python dataclasses."""

        logging.info("Preparing Workbook Data...")

        # The following columns have some invisible formatting that ultimately
        # resolve to None, with no discernible value, which is causing errors
        # further down the pipeline, so we can terminate them silently.
        logging.info("Unnecessary metadata detected -- Removing column 'P' from Invoices...")
        self._inv.delete_cols(16)  # Column P
        logging.info("Unnecessary metadata detected -- Removing column 'G' from Direct Payroll Adjustments...")
        self._dpa.delete_cols(7)  # Column G

        # We can also reliably delete the following sheets from the result workbook.
        if "Commission Base Payroll Adj" in self.workbook.sheetnames:
            logging.info("Unnecessary sheet detected -- Removing sheets 'Commission Base Payroll Adj'...")
            del self.workbook["Commission Base Payroll Adj"]
        if "Non-job Purchase Orders" in self.workbook.sheetnames:
            logging.info("Unnecessary sheet detected -- Removing sheets 'Non-job Purchase Orders'...")
            del self.workbook["Non-job Purchase Orders"]

        # Populate Invoices
        inv_header, dpa_header = None, None
        for i, row in enumerate(self._inv.iter_rows()):
            if i == 0:
                # Edit the following to make the header row viable as dataclass keys
                # - Remove trailing percentage sign: Split % -> split
                # - Replace spaces with underscores: Invoiced On -> invoiced_on
                # - Lowercase all: Business Unit -> business_unit
                # - Remove periods: Pay Adj. -> pay_adj
                inv_header = [cell.value.lower().rstrip(" %").replace(".", "").replace(" ", "_") for cell in row]
                continue

            # If there was no header, invalid format
            if inv_header is None:
                return

            # Create a temporary dict to hold the data and assign each element to the appropriate key
            data = {}
            for j, cell in enumerate(row):
                data[inv_header[j]] = cell.value  # type: ignore

            # Then unpack into a more manageable Python object and append to the technician's collection
            inv = Invoice(**data)
            self._technician_dict[inv.technician]["invoices"].append(inv)

        # Create some lookup tables so we can cross-reference later
        logging.info("Invoices populated...")

        # Populate Direct Payroll Adjustments
        for i, row in enumerate(self._dpa.iter_rows()):
            # Ignore header
            if row[0].row == 1:
                dpa_header = [cell.value.lower().rstrip(" %").replace(".", "").replace(" ", "_") for cell in row]
                continue

            # If there was no header, invalid format
            if dpa_header is None:
                return

            # Assign data to its appropriate key
            data = {}
            for j, cell in enumerate(row):
                data[dpa_header[j]] = cell.value  # type: ignore

            dpa = DirectPayrollAdjustment(*[cell.value for cell in row])
            self._technician_dict[dpa.technician]["dpa"].append(dpa)

        logging.info(f"Direct Payroll Adjustments populated: {len(self._technician_dict)} total technicians found in all.")

################################################################################
    def run_merge(self) -> bool:
        """Run the merge operation to create technician-specific sheets."""

        # Start off with the master sheet that all records will be added to
        self._master_tech_sheet = self.workbook.create_sheet("All Technicians", index=2)

        # Resize columns to fit data - can't auto-fit like Excel, so we set manually
        # These are measured widths (Column Width from context menu)
        column_widths = [10, 15, 25, 20, 15, 10, 15, 15, 8, 18, 15, 22]
        for i, column_width in enumerate(column_widths, start=1):
            self._master_tech_sheet.column_dimensions[chr(64 + i)].width = column_width

        # Helper function to determine if a technician sheet needs to be created.
        def needs_to_be_created(name: str, data: Dict[str, List[Union[Invoice, DirectPayrollAdjustment]]]) -> bool:
            return (
                # Tech name must not be None
                name is not None
                # There must be at least one invoice or DPA for this tech
                and (len(data["invoices"]) > 0 or len(data["dpa"]) > 0)
                # There must be at least one invoice with GP > $0.00 or DPA with any amount
                and (sum(inv.gp for inv in data["invoices"]) > 0.0 or len(data["dpa"]) > 0)
            )

        # Now iterate through each technician and create their sheets
        for tech_name, values in self._technician_dict.items():
            # Tech name is `None` on totals row at the bottom. Skip it.
            # Also skip processing if there are no eligible records in this cycle for this tech.
            if not needs_to_be_created(tech_name, values):
                continue

            # Create a sheet for this technician
            self.workbook.create_sheet(tech_name)
            # Populate the sheet and capture any records missing customer names
            self._populate_tech_data(tech_name)

        # Add Conditional Formatting to highlight missing customer names in the master sheet
        self._master_tech_sheet.conditional_formatting.add(f"C3:C{self.master_current_row}", IS_BLANK_RULE)
        # Make sure number-related columns (E:L) are formatted as currency
        for col in self._master_tech_sheet.iter_cols(min_col=5, max_col=12, min_row=3, max_row=self._master_tech_sheet.max_row):
            for cell in col:
                cell.number_format = FORMAT_CURRENCY_USD_SIMPLE

        # Simple flag to indicate success
        flag = True

        # Dear god, don't forget to save the workbook! T_T
        # I/O operations can fail for any number of reasons, so wrap in try/except
        try:
            wb_name = f"{self._wb_path.stem}-Combined.xlsx"
            self.out_wb_path = self.output_dir / wb_name
            logging.info(f"Saving combined workbook to {self.out_wb_path}")
            self.workbook.save(self.out_wb_path)
            logging.info("Workbook saved...")
        except PermissionError:
            logging.critical(f"Permission denied when saving workbook (is it open in another program?)")
            # Don't return here, we still want to attempt cleanup
            flag = False
        except Exception as ex:
            logging.critical(f"Error saving workbook: {ex}")
            # Don't return here, we still want to attempt cleanup
            flag = False

        if self._lookup_path.exists():
            logging.debug(f"Removing temporary lookup file at {self._lookup_path}...")
            try:
                self._lookup_path.unlink(missing_ok=True)
            except Exception as ex:
                logging.error(f"Failed to remove temporary lookup file: {ex}")

        return flag

################################################################################
    def _populate_tech_data(self, tech_name: str) -> None:
        """Populate the given technician's sheet with their specific data."""

        # Ensure the sheet exists (it should, we just created it, hence the assertion)
        logging.info(f"Preparing report for technician: {tech_name}...")
        sheet: Worksheet = self.workbook[tech_name]  # type: ignore

        # We need to keep a counter of the current row for positioning
        self.current_row = 1

        # Reusable local function to create header rows
        def create_header_row(text: str, b: bool = True, i: bool = False, size: int = 11) -> None:
            # Merge the data columns together
            sheet.merge_cells(f"A{self.current_row}:E{self.current_row}")
            self._master_tech_sheet.merge_cells(f"A{self.master_current_row}:E{self.master_current_row}")
            header_row: Cell = sheet[f"A{self.current_row}"]  # type: ignore
            master_header_row: Cell = self._master_tech_sheet[f"A{self.master_current_row}"]  # type: ignore
            # Center align and apply font styling
            header_row.alignment = Alignment(horizontal="center", vertical="center")
            header_row.font = Font(size=size, bold=b, italic=i)
            master_header_row.alignment = Alignment(horizontal="center", vertical="center")
            master_header_row.font = Font(size=size, bold=b, italic=i)
            # Set the header text
            header_row.value = text
            master_header_row.value = text
            # Increment the row counters
            self.current_row += 1
            self.master_current_row += 1

        def _get_cell_ref(c: str) -> str:
            """Helper function to get a reference to a cell in this technician's workbook."""
            ref = f"'{tech_name}'!{c}"
            # If the cell is blank, return an empty string, else return the cell reference. If we
            # don't do this, Excel will display 0 for blank cells which is incompatible with
            # the conditional formatting.
            return f"=IF({ref}=\"\",\"\",{ref})"

        # Local function to append a row and increment the counter
        def append_row(values: List[Union[str, float]], as_refs_to_master: bool = False) -> None:
            sheet.append(values)
            if not as_refs_to_master:
                self._master_tech_sheet.append(values)
            else:
                # If it's a reference, we need to adjust the formula to point to the correct cell(s)
                mrow = list(self._master_tech_sheet.iter_rows(min_row=self.master_current_row, max_row=self.master_current_row, min_col=1, max_col=5))[0]  # type: ignore
                for i, c in enumerate(mrow):
                    c.value = _get_cell_ref(f"{get_column_letter(i + 1)}{self.current_row}")
            self.current_row += 1
            self.master_current_row += 1

        # First handle overall sheet setup and formatting
        # Add tech name as a title header
        create_header_row(tech_name, size=14)
        # Add data labels
        header_items = ["Invoice", "Invoiced On", "Customer Name", "Memo", "Amount"]
        sheet.append(header_items)
        self.current_row += 1
        master_header_items = ["Invoice", "Invoiced On", "Customer Name", "Memo", "Amount", "Subtotal", "Commission (S)", "Commission (I)", "SPIFFS", "Sales Commission", "Truck Revenue", "Marketing Commission"]
        self._master_tech_sheet.append(master_header_items)
        self.master_current_row += 1
        # Format label row (Bold & centered) (Append row auto-augments the current row
        # so ensure we subtract 1 to get the correct label row)
        label_row_tech: List[Cell] = list(sheet.iter_rows(min_row=self.current_row - 1, max_row=self.current_row - 1, min_col=1, max_col=len(header_items)))[0]  # type: ignore
        label_row_master: List[Cell] = list(self._master_tech_sheet.iter_rows(min_row=self.master_current_row - 1, max_row=self.master_current_row - 1, min_col=1, max_col=len(master_header_items)))[0]  # type: ignore
        for cell in label_row_tech:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="left" if cell.column != 5 else "right", vertical="center")
        for cell in label_row_master:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Resize columns to fit data - can't auto-fit like Excel, so we set manually
        column_widths = [10, 15, 25, 20, 15]  # These are measured widths (Column Width from context menu)
        for i, column_width in enumerate(column_widths, start=1):
            sheet.column_dimensions[chr(64 + i)].width = column_width

        # Isolate this technician's data from the master dict
        tech_invoices = self._technician_dict[tech_name]["invoices"]
        tech_dpas = self._technician_dict[tech_name]["dpa"]
        combined = tech_invoices + tech_dpas
        # Sort combined list by date (invoiced_on for invoices, posted_on for DPAs)
        items_by_date = sorted(combined, key=lambda item: item.invoiced_on if isinstance(item, Invoice) else item.posted_on)

        # Create a reference lookup for customer names by invoice ID from the invoices only
        invoice_customer_lookup = {inv.invoice: inv.customer for inv in tech_invoices}
        amount_total = 0.0

        for inv in items_by_date:
            # If there's a matching DPA, include the memo and amount in the invoice section
            dt = inv.invoiced_on if isinstance(inv, Invoice) else inv.posted_on
            memo = inv.memo if isinstance(inv, DirectPayrollAdjustment) else "---"
            amount = inv.amount if isinstance(inv, DirectPayrollAdjustment) else inv.gp
            amount_total += amount
            customer = invoice_customer_lookup.get(inv.invoice, None)

            # If the amount is $0.00 skip it, tech doesn't need to see it
            # Don't skip $0.00 DPAs though, those are important to show
            if amount == 0.0 and isinstance(inv, Invoice):
                continue

            # If customer name is missing, prompt user to enter it
            if customer is None:
                customer = self._lookup_table.get(inv.invoice, "")
                if customer == "":
                    logging.warning(f"Missing customer name for invoice {inv.invoice} for technician {tech_name}.")

            # Append the row to the sheet
            append_row([inv.invoice, dt.strftime("%m/%d/%Y"), customer, memo, amount], True)

        # Add a total row at the bottom
        sheet.append(["", "", "", "Total:", f"=SUM(E3:E{self.current_row - 1})"])
        self.current_row += 1
        self._master_tech_sheet.append(["", "", "", "", "Total:", _get_cell_ref(f"E{self.current_row - 1}")])
        self.master_current_row += 1

        tech_total_row: List[Cell] = list(sheet.iter_rows(min_row=self.current_row - 1, max_row=self.current_row - 1, min_col=1, max_col=5))[0]  # type: ignore
        master_total_row: List[Cell] = list(self._master_tech_sheet.iter_rows(min_row=self.master_current_row - 1, max_row=self.master_current_row - 1, min_col=1, max_col=5))[0]  # type: ignore
        # Format 'TOTAL' rows in both sheets (Bold, right-aligned on Amount)
        for cell in tech_total_row:
            cell.font = Font(bold=True)
            if cell.column == 4:  # Amount column
                cell.alignment = Alignment(horizontal="right", vertical="center")
        for cell in master_total_row[4:6]:  # Only Total label and value columns
            cell.font = Font(bold=True)
            if cell.column == 5:  # Amount column
                cell.alignment = Alignment(horizontal="right", vertical="center")

        # Apply conditional formatting for the Custom Name column to highlight blanks
        # We have to do this after all data is populated, so we know the range
        # starting at row 2 (below the header) to the current row - 1 (total row)
        if self.current_row > 3:  # Only apply if there's data rows
            sheet.conditional_formatting.add(f"C3:C{self.current_row - 1}", IS_BLANK_RULE)

        # Make sure the Amount column is formatted as currency
        for cell in sheet["E"]:  # type: ignore
            cell.number_format = FORMAT_CURRENCY_USD_SIMPLE

        # Apply borders to all rows aside from total row
        for row in sheet.iter_rows(min_row=1, max_row=self.current_row - 2, min_col=1, max_col=5):
            for cell in row:
                cell.border = ROW_BORDER

        self.autofit_column(sheet, self.current_row)

################################################################################
    def export_pdfs(self, print_date: str) -> None:
        """Export each technician's sheet as a separate PDF file."""

        # PDF export is only supported on Windows systems with Excel installed
        if not IS_WINDOWS:
            logging.critical("PDF export is only supported on Windows systems with Excel installed.")
            return

        logging.info(f"Extracting PDF files...")

        # We need a handle to the Excel COM object to give it direct commands.
        excel_dispatch = win32.Dispatch("Excel.Application")
        excel_dispatch.Visible = False
        excel_dispatch.DisplayAlerts = False

        # Ensure output directory exists
        out_path = Path(self.output_dir) / "PDFs"
        out_path.mkdir(parents=True, exist_ok=True)

        # We don't want to export the following, only individual technician sheets
        sheets_to_skip = ["Invoices", "Direct Payroll Adjustments", "All Technicians"]
        # now = datetime.now()
        # date_slug = f"{now.month}-{now.day}-{now.year}"

        wb = None
        try:
            # Open the workbook via COM
            wb = excel_dispatch.Workbooks.Open(str(self.out_wb_path))
            for ws in wb.Worksheets:
                # Skip non-technician sheets
                if ws.Name in sheets_to_skip:
                    continue

                # Get the technician's name and format the PDF filename
                fn, ln = ws.Name.split(" ", 1) if " " in ws.Name else (ws.Name, "")
                pdf_slug = f"{ln}, {fn} - {print_date}.pdf"
                pdf_path = out_path / pdf_slug

                # Export the worksheet as a PDF scaled to fit a single page wide
                # Don't try to move this to the end of `_populate_tech_data`. It
                # doesn't work properly there for some reason. You've tried.
                ws.PageSetup.Zoom = False
                ws.PageSetup.FitToPagesWide = 1
                ws.PageSetup.FitToPagesTall = False  # Unlimited length
                try:
                    logging.info(f"Extracting {ws.Name}...")
                    ws.ExportAsFixedFormat(Type=0, Filename=str(pdf_path), Quality=0, IncludeDocProperties=True, IgnorePrintAreas=False)
                except com_error:
                    logging.critical(f"Failed to export PDF for technician {ws.Name}: Is the PDF file already open?")
                    continue
        except Exception as ex:
            logging.critical(f"Failed to open workbook for PDF export: {ex}")
            return
        finally:
            # Clean up COM objects
            if wb is not None:
                wb.Close(False)
            excel_dispatch.Quit()

################################################################################
    @staticmethod
    def _as_display_text(value: Any, num_fmt: Optional[str]) -> str:
        """Best effort attempt at text width estimation for auto-fitting columns."""

        # If there's no value, we can return an empty string
        if value is None:
            return ""

        # This is relatively crude since real formatting is complex and relies on number_fmt/locale
        if isinstance(value, (date, datetime, time)):
            if isinstance(value, datetime):
                return value.strftime("%m/%d/%Y %H:%M")
            elif isinstance(value, date):
                return value.strftime("%m/%d/%Y")
            else:
                return value.strftime("%H:%M")

        # Currency detection
        if isinstance(value, (int, float)) and num_fmt is not None:
            nf = num_fmt.replace(" ", "").lower()
            if "0.00" in nf and ("$" in nf or "usd" in nf):
                # Add separators and 2 decimal places for width
                return f"${value:,.2f}"

        # Fallback to string
        return str(value)

################################################################################
    def autofit_column(
        self,
        ws: Worksheet,
        end_row: int,
        *,
        min_width: float = 15.0,
        max_width: float = 35.0,
        padding: float = 2.0,
        bold_scalar: float = 1.08
    ) -> None:
        """Auto-fits the given column in `ws` based on estimated content width."""

        max_chars = 0.0
        # Cycle through each cell in the column to determine the maximum text width
        for row in ws.iter_rows(min_row=2, max_row=end_row, min_col=4, max_col=4):
            cell: Cell = row[0]
            txt = self._as_display_text(cell.value, getattr(cell, "number_format", None))
            if not txt:
                continue

            # This handles multi-line comments by taking the widest line
            widest_line = max(len(line) for line in str(txt).splitlines())
            # If there's bold formatting, scale the width accordingly
            if cell.font and cell.font.bold:
                widest_line *= bold_scalar
            # If the widest line exceeds the current max, update it
            if widest_line > max_chars:
                max_chars = widest_line

            # Finally, enable line wrapping to cover any memos that exceed the column width
            # Center-justify header row, left-align others
            cell.alignment = Alignment(wrap_text=True, horizontal="left")

        # Finally, set the column width with padding, clamped to min/max
        # so it will still fit within the bounds of a page when printed
        width = min(max_width, max(min_width, max_chars + padding))
        ws.column_dimensions["D"].width = width

################################################################################
